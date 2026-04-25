import csv
import io
import os

from app.bob50_service import generer_fichiers_bob50
from datetime import datetime
from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file, Response

from werkzeug.datastructures import ImmutableMultiDict
from sqlalchemy import func, case

from app.extensions import db
from app.models import (
    ParentEleve, Eleve, Activite, ActiviteEleve, PaiementActivite, MouvementPortefeuille,TypeRepas, CongeRepas, ReservationRepas
)
from app.qr_service import generer_qr_code_base64, generer_qr_code_png

from app.forms import ActiviteForm    # Formulaire WTForms pour la création d'activité

main = Blueprint("main", __name__)


# ─────────────────────────────────────────────────────────────
# TABLEAU DE BORD ADMIN
# ─────────────────────────────────────────────────────────────

@main.route("/admin")
def admin_dashboard():
    """Tableau de bord avec statistiques globales."""

    nb_activites_ouvertes = (
        db.session.query(func.count(Activite.id))
        .filter(Activite.statut == "ouvert")
        .scalar()
    )

    nb_en_attente = (
        db.session.query(func.count(ActiviteEleve.id))
        .filter(ActiviteEleve.statut == "en_attente")
        .scalar()
    )

    row = (
        db.session.query(
            func.count(PaiementActivite.id),
            func.coalesce(func.sum(PaiementActivite.montant), 0)
        )
        .filter(PaiementActivite.statut == "paye")
        .filter(
            func.date_trunc("month", PaiementActivite.paye_le)
            == func.date_trunc("month", func.now())
        )
        .first()
    )

    stats = {
        "nb_activites_ouvertes": nb_activites_ouvertes,
        "nb_en_attente": nb_en_attente,
        "nb_payes": row[0],
        "montant_encaisse": f"{float(row[1]):.2f}",
    }

    return render_template(
        "admin_dashboard.html",
        stats=stats,
        role="admin",
        active="dashboard",
        current_user_name="Administrateur (test)"
    )


# ─────────────────────────────────────────────────────────────
# ESPACE PARENT
# ─────────────────────────────────────────────────────────────

@main.route("/")
def index():
    """Liste des activités liées aux enfants du parent connecté."""

    parent_id = 1  # TODO : flask_login.current_user.id

    lignes = (
        ActiviteEleve.query
        .join(Eleve, ActiviteEleve.eleve_id == Eleve.id)
        .join(Activite, ActiviteEleve.activite_id == Activite.id)
        .join(ParentEleve, ParentEleve.eleve_id == Eleve.id)
        .filter(ParentEleve.parent_id == parent_id)
        .order_by(Eleve.nom, Eleve.prenom)
        .all()
    )

    return render_template(
        "index.html",
        lignes=lignes,
        role="parent",
        active="accueil",
        current_user_name="Parent (test)"
    )


# ─────────────────────────────────────────────────────────────
# QR CODE — PAIEMENT PAR VIREMENT
# ─────────────────────────────────────────────────────────────

@main.route("/mon-qrcode/<int:activite_eleve_id>")
def mon_qrcode(activite_eleve_id):
    """
    Page QR code EPC + instructions de virement pour le parent.
    Remplace l'ancienne page confirmation_paiement.
    """

    parent_id = 1  # TODO : flask_login.current_user.id

    # Sécurité : vérification que l'activité appartient à un enfant du parent
    ae = (
        ActiviteEleve.query
        .join(Eleve, ActiviteEleve.eleve_id == Eleve.id)
        .join(ParentEleve, ParentEleve.eleve_id == Eleve.id)
        .filter(ActiviteEleve.id == activite_eleve_id)
        .filter(ParentEleve.parent_id == parent_id)
        .first()
    )

    if ae is None:
        flash("Activité introuvable ou accès non autorisé.", "erreur")
        return redirect(url_for("main.index"))

    if ae.statut == "paye":
        flash("Cette activité est déjà payée.", "info")
        return redirect(url_for("main.index"))

    # Génération du QR code en base64 pour affichage inline dans le template
    qr_b64, reference = generer_qr_code_base64(
        activite_id=ae.activite_id,
        eleve_id=ae.eleve_id,
        montant=float(ae.montant_attendu),
        description=ae.activite.titre
    )

    return render_template(
        "mon_qrcode.html",
        ae=ae,
        qr_b64=qr_b64,
        reference=reference,
        ecole_nom=os.getenv("NOM_BENEFICIAIRE", ""),
        ecole_iban=os.getenv("IBAN", ""),
        role="parent",
        active="accueil",
        current_user_name="Parent (test)"
    )


@main.route("/qrcode-image/<int:activite_eleve_id>")
def qrcode_image(activite_eleve_id):
    """
    Retourne l'image PNG du QR code pour téléchargement direct.
    Utilisé par le bouton "Télécharger" dans mon_qrcode.html.
    """

    parent_id = 1  # TODO : flask_login.current_user.id

    ae = (
        ActiviteEleve.query
        .join(Eleve, ActiviteEleve.eleve_id == Eleve.id)
        .join(ParentEleve, ParentEleve.eleve_id == Eleve.id)
        .filter(ActiviteEleve.id == activite_eleve_id)
        .filter(ParentEleve.parent_id == parent_id)
        .first()
    )

    if ae is None:
        return "Non autorisé", 403

    png_bytes, _ = generer_qr_code_png(
        activite_id=ae.activite_id,
        eleve_id=ae.eleve_id,
        montant=float(ae.montant_attendu),
        description=ae.activite.titre
    )

    return send_file(
        io.BytesIO(png_bytes),
        mimetype="image/png",
        download_name=f"paiement_{activite_eleve_id}.png"
    )


# ─────────────────────────────────────────────────────────────
# BACK-OFFICE ADMIN — PAIEMENTS
# ─────────────────────────────────────────────────────────────

@main.route("/admin/paiements")
def admin_paiements():
    """Liste de toutes les activités-élèves avec leur statut de paiement."""

    lignes = (
        ActiviteEleve.query
        .join(Eleve, ActiviteEleve.eleve_id == Eleve.id)
        .join(Activite, ActiviteEleve.activite_id == Activite.id)
        .order_by(ActiviteEleve.statut, Eleve.nom, Eleve.prenom)
        .all()
    )

    return render_template(
        "admin_paiements.html",
        lignes=lignes,
        role="admin",
        active="paiements",
        current_user_name="Administrateur (test)"
    )


@main.route("/admin/marquer-paye/<int:activite_eleve_id>", methods=["POST"])
def admin_marquer_paye(activite_eleve_id):
    """Valide manuellement un virement bancaire reçu par le secrétariat."""

    ligne = ActiviteEleve.query.get(activite_eleve_id)

    if ligne is None:
        flash("Activité introuvable.", "erreur")
        return redirect(url_for("main.admin_paiements"))

    if ligne.statut == "paye":
        flash("Cette activité est déjà marquée comme payée.", "info")
        return redirect(url_for("main.admin_paiements"))

    paiement = PaiementActivite(
        activite_eleve_id=ligne.id,
        montant=ligne.montant_attendu,
        statut="paye",
        mode_paiement="virement",
        ref_transaction=f"VIR-{ligne.id}",
        paye_le=db.func.now()
    )
    db.session.add(paiement)
    ligne.statut = "paye"
    db.session.commit()

    flash("Virement validé et paiement enregistré.", "succes")
    return redirect(url_for("main.admin_paiements"))


# ─────────────────────────────────────────────────────────────
# BACK-OFFICE ADMIN — ACTIVITÉS
# ─────────────────────────────────────────────────────────────

@main.route("/admin/activites")
def admin_activites():
    """Liste des activités avec comptage des paiements reçus."""

    activites = (
        db.session.query(
            Activite.id,
            Activite.titre,
            Activite.montant,
            Activite.date_limite_paiement,
            Activite.statut,
            func.count(ActiviteEleve.id).label("nb_total"),
            func.count(
                case((ActiviteEleve.statut == "paye", 1))
            ).label("nb_payes")
        )
        .outerjoin(ActiviteEleve, ActiviteEleve.activite_id == Activite.id)
        .group_by(Activite.id)
        .order_by(Activite.date_limite_paiement.desc())
        .all()
    )

    return render_template(
        "admin_activites.html",
        activites=activites,
        role="admin",
        active="activites",
        current_user_name="Administrateur (test)"
    )


@main.route("/admin/activites/nouvelle", methods=["GET", "POST"])
def admin_nouvelle_activite():
    """
    GET  : affiche le formulaire de création d'activité
    POST : valide via WTForms, crée l'activité et affecte les élèves
    """

    # Récupération des classes actives pour les cases à cocher
    classes_dispo = [
        row[0]
        for row in (
            db.session.query(Eleve.classe)
            .filter(Eleve.actif == True)
            .distinct()
            .order_by(Eleve.classe)
            .all()
        )
    ]

    form = ActiviteForm()

    if form.validate_on_submit():
        # Récupération des classes cochées (pas géré par WTForms car liste dynamique)
        classes_cibles = request.form.getlist("classes")

        if not classes_cibles:
            flash("Sélectionnez au moins une classe.", "erreur")
            return render_template(
                "admin_nouvelle_activite.html",
                classes_dispo=classes_dispo,
                form=form,
                role="admin",
                active="nouvelle_activite",
                current_user_name="Administrateur (test)"
            )

        # Création de l'activité avec les données validées par WTForms
        activite = Activite(
            titre=form.titre.data,
            description=form.description.data or None,
            montant=form.montant.data,
            date_limite_paiement=form.date_limite_paiement.data,
            obligatoire=form.obligatoire.data,
            statut="ouvert"
        )
        db.session.add(activite)
        db.session.flush()  # Récupère l'id sans commit

        # Affectation aux élèves des classes sélectionnées
        eleves = (
            Eleve.query
            .filter(Eleve.classe.in_(classes_cibles))
            .filter(Eleve.actif == True)
            .all()
        )

        nb_affectes = 0
        for eleve in eleves:
            existe = ActiviteEleve.query.filter_by(
                activite_id=activite.id,
                eleve_id=eleve.id
            ).first()
            if not existe:
                ae = ActiviteEleve(
                    activite_id=activite.id,
                    eleve_id=eleve.id,
                    montant_attendu=form.montant.data,
                    statut="en_attente"
                )
                db.session.add(ae)
                nb_affectes += 1

        db.session.commit()

        flash(
            f"Activité « {form.titre.data} » créée et affectée à "
            f"{nb_affectes} élève(s) dans les classes : "
            f"{', '.join(classes_cibles)}.",
            "succes"
        )
        return redirect(url_for("main.admin_activites"))

    # GET ou formulaire invalide : affichage du formulaire
    return render_template(
        "admin_nouvelle_activite.html",
        classes_dispo=classes_dispo,
        form=form,
        role="admin",
        active="nouvelle_activite",
        current_user_name="Administrateur (test)"
    )

# ─────────────────────────────────────────────────────────────
# IMPORT CSV / XLSX ÉLÈVES
# ─────────────────────────────────────────────────────────────

@main.route("/admin/eleves/importer", methods=["GET", "POST"])
def admin_importer_eleves():
    """
    Import des élèves en début d'année scolaire.
    Supporte deux formats :
    - ProEco (.xlsx) : colonnes Nom Elève, Prénom Elève, Année, Classe, etc.
    - CSV manuel    : colonnes prenom, nom, classe, option, annee_scolaire
    """

    ctx = {
        "role": "admin",
        "active": "import_eleves",
        "current_user_name": "Administrateur (test)"
    }

    if request.method == "GET":
        return render_template("admin_importer_eleves.html", **ctx)

    fichier = request.files.get("fichier_csv")

    if not fichier or fichier.filename == "":
        flash("Aucun fichier sélectionné.", "erreur")
        return render_template("admin_importer_eleves.html", **ctx)

    nb_inseres    = 0
    nb_mis_a_jour = 0
    nb_erreurs    = 0
    erreurs_detail = []

    # ── Format ProEco (.xlsx) ─────────────────────────────────
    if fichier.filename.lower().endswith(".xlsx"):

        # Année scolaire obligatoire pour le format ProEco
        annee_scolaire = request.form.get("annee_scolaire", "").strip()
        if not annee_scolaire:
            flash("L'année scolaire est obligatoire pour l'import ProEco.", "erreur")
            return render_template("admin_importer_eleves.html", **ctx)

        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(fichier.read()))
            ws = wb.active

            # Récupère les en-têtes de la première ligne
            headers = [cell.value for cell in ws[1]]

            # Vérifie que les colonnes requises sont présentes
            colonnes_requises = {'Nom Elève', 'Prénom Elève', 'Année', 'Classe', 'Matric Info'}
            if not colonnes_requises.issubset(set(headers)):
                manquantes = colonnes_requises - set(headers)
                flash(f"Colonnes manquantes : {', '.join(manquantes)}", "erreur")
                return render_template("admin_importer_eleves.html", **ctx)

            # Construit un dictionnaire {nom_colonne: index} pour accès rapide
            idx = {h: i for i, h in enumerate(headers)}

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

                # Ignore les lignes vides
                if not any(row):
                    continue

                nom    = str(row[idx['Nom Elève']]    or '').strip()
                prenom = str(row[idx['Prénom Elève']] or '').strip()
                annee  = str(row[idx['Année']]         or '').strip()
                classe_section = str(row[idx['Classe']] or '').strip()
                matricule = str(row[idx['Matric Info']] or '').strip() or None

                # Option = Sous Groupe + Langue si disponibles
                # Ex : "SG - A" (Sciences Général - Anglais)
                sous_groupe = str(row[idx['Sous Groupe']] or '').strip() if 'Sous Groupe' in idx else ''
                langue      = str(row[idx['Langue I  2e langue ']] or '').strip() if 'Langue I  2e langue ' in idx else ''

                if sous_groupe and langue:
                    option = f"{sous_groupe} - {langue}"
                elif sous_groupe:
                    option = sous_groupe
                elif langue:
                    option = langue
                else:
                    option = None

                # Classe = Année + espace + Section → ex: "3 A" ou "2C A"
                classe = f"{annee} {classe_section}".strip() if classe_section else annee

                if not all([nom, prenom, classe]):
                    nb_erreurs += 1
                    erreurs_detail.append(f"Ligne {i} ignorée — champs manquants")
                    continue

                eleve = Eleve.query.filter_by(
                    prenom=prenom,
                    nom=nom,
                    classe=classe,
                    annee_scolaire=annee_scolaire
                ).first()

                if eleve is None:
                    eleve = Eleve(
                        prenom=prenom,
                        nom=nom,
                        classe=classe,
                        option=option,
                        annee_scolaire=annee_scolaire,
                        matricule_fase=matricule,
                        actif=True
                    )
                    db.session.add(eleve)
                    nb_inseres += 1
                else:
                    eleve.option = option
                    eleve.actif = True
                    eleve.matricule_fase = matricule
                    nb_mis_a_jour += 1

            db.session.commit()

        except Exception as e:
            db.session.rollback()
            flash(f"Erreur lors de l'import ProEco : {str(e)}", "erreur")
            return render_template("admin_importer_eleves.html", **ctx)

    # ── Format CSV manuel ─────────────────────────────────────
    elif fichier.filename.lower().endswith(".csv"):

        try:
            contenu = fichier.read().decode("utf-8-sig")
            lecteur = csv.DictReader(io.StringIO(contenu))

            # Vérifie les colonnes requises
            colonnes_requises = {"prenom", "nom", "classe", "annee_scolaire"}
            if not colonnes_requises.issubset(set(lecteur.fieldnames or [])):
                manquantes = colonnes_requises - set(lecteur.fieldnames or [])
                flash(
                    f"Colonnes manquantes : {', '.join(manquantes)}. "
                    "Colonnes attendues : prenom, nom, classe, option, annee_scolaire",
                    "erreur"
                )
                return render_template("admin_importer_eleves.html", **ctx)

            lignes = list(lecteur)

        except UnicodeDecodeError:
            flash(
                "Erreur d'encodage. Enregistrez le fichier en UTF-8 "
                "(Excel : Enregistrer sous → CSV UTF-8).",
                "erreur"
            )
            return render_template("admin_importer_eleves.html", **ctx)

        if not lignes:
            flash("Le fichier CSV est vide.", "erreur")
            return render_template("admin_importer_eleves.html", **ctx)

        try:
            for i, ligne in enumerate(lignes, start=2):
                prenom        = ligne.get("prenom", "").strip()
                nom           = ligne.get("nom", "").strip()
                classe        = ligne.get("classe", "").strip()
                option        = ligne.get("option", "").strip() or None
                annee_scolaire = ligne.get("annee_scolaire", "").strip()
                matricule_fase = ligne.get("matricule_fase", "").strip() or None

                if not all([prenom, nom, classe, annee_scolaire]):
                    nb_erreurs += 1
                    erreurs_detail.append(
                        f"Ligne {i} ignorée — champs obligatoires manquants : "
                        f"{prenom!r}, {nom!r}, {classe!r}, {annee_scolaire!r}"
                    )
                    continue

                eleve = Eleve.query.filter_by(
                    prenom=prenom,
                    nom=nom,
                    classe=classe,
                    annee_scolaire=annee_scolaire
                ).first()

                if eleve is None:
                    eleve = Eleve(
                        prenom=prenom,
                        nom=nom,
                        classe=classe,
                        option=option,
                        annee_scolaire=annee_scolaire,
                        matricule_fase=matricule_fase,
                        actif=True
                    )
                    db.session.add(eleve)
                    nb_inseres += 1
                else:
                    eleve.option = option
                    eleve.actif = True
                    eleve.matricule_fase = matricule_fase
                    nb_mis_a_jour += 1

            db.session.commit()

        except Exception as e:
            db.session.rollback()
            flash(f"Erreur lors de l'import : {str(e)}", "erreur")
            return render_template("admin_importer_eleves.html", **ctx)

    else:
        # Format non supporté
        flash("Format non supporté. Utilisez un fichier .xlsx (ProEco) ou .csv.", "erreur")
        return render_template("admin_importer_eleves.html", **ctx)

    # ── Messages de résultat ──────────────────────────────────
    if nb_inseres > 0 or nb_mis_a_jour > 0:
        flash(
            f"Import terminé : {nb_inseres} élève(s) ajouté(s), "
            f"{nb_mis_a_jour} élève(s) mis à jour.",
            "succes"
        )
    if nb_erreurs > 0:
        flash(f"{nb_erreurs} ligne(s) ignorée(s).", "erreur")
        for detail in erreurs_detail[:5]:
            flash(detail, "erreur")

    return redirect(url_for("main.admin_importer_eleves"))
# ─────────────────────────────────────────────────────────────
# EXPORT PAIEMENTS — À ajouter dans routes.py
# ─────────────────────────────────────────────────────────────
# Ajouter en haut de routes.py si pas déjà présent :
#   import csv
#   import io
#   from datetime import datetime
#   from flask import Response


@main.route("/admin/export")
def admin_export():
    """
    Page d'export : affiche les filtres et un aperçu des résultats.
    Les filtres sont passés via les paramètres GET de l'URL.
    """

    # ── Récupération des filtres depuis l'URL ────────────────
    activite_id  = request.args.get("activite_id", "")
    classe       = request.args.get("classe", "")
    statut       = request.args.get("statut", "paye")   # défaut : payés
    date_debut   = request.args.get("date_debut", "")
    date_fin     = request.args.get("date_fin", "")

    filtres = {
        "activite_id": activite_id,
        "classe":      classe,
        "statut":      statut,
        "date_debut":  date_debut,
        "date_fin":    date_fin,
    }

    # ── Données pour les menus déroulants ────────────────────
    activites_dispo = (
        Activite.query
        .order_by(Activite.date_limite_paiement.desc())
        .all()
    )

    classes_dispo = [
        row[0]
        for row in (
            db.session.query(Eleve.classe)
            .filter(Eleve.actif == True)
            .distinct()
            .order_by(Eleve.classe)
            .all()
        )
    ]

    # ── Construction de la requête avec les filtres ──────────
    lignes = _construire_requete_export(
        activite_id, classe, statut, date_debut, date_fin
    )

    return render_template(
        "admin_export.html",
        lignes=lignes,
        activites_dispo=activites_dispo,
        classes_dispo=classes_dispo,
        filtres=filtres,
        role="admin",
        active="export",
        current_user_name="Administrateur (test)"
    )


@main.route("/admin/export/csv")
def admin_export_csv():
    """
    Génère et télécharge directement le fichier CSV.
    Utilise les mêmes filtres que la page d'aperçu.
    """

    activite_id  = request.args.get("activite_id", "")
    classe       = request.args.get("classe", "")
    statut       = request.args.get("statut", "paye")
    date_debut   = request.args.get("date_debut", "")
    date_fin     = request.args.get("date_fin", "")

    lignes = _construire_requete_export(
        activite_id, classe, statut, date_debut, date_fin
    )

    # ── Génération du CSV en mémoire ─────────────────────────
    output = io.StringIO()
    writer = csv.writer(output, delimiter=",", quoting=csv.QUOTE_MINIMAL)

    # En-tête
    writer.writerow([
        "activite",
        "classe",
        "nom_eleve",
        "prenom_eleve",
        "option",
        "statut",
        "montant",
        "reference",
        "date_validation",
        "mode_paiement",
    ])

    # Données
    for l in lignes:
        # Récupère le paiement validé s'il existe
        paiement = l.paiements[0] if l.paiements else None

        writer.writerow([
            l.activite.titre,
            l.eleve.classe,
            l.eleve.nom,
            l.eleve.prenom,
            l.eleve.option or "",
            l.statut,
            str(l.montant_attendu),
            paiement.ref_transaction if paiement else "",
            paiement.paye_le.strftime("%d/%m/%Y %H:%M") if (paiement and paiement.paye_le) else "",
            paiement.mode_paiement if paiement else "",
        ])

    # ── Nom du fichier avec date du jour ─────────────────────
    nom_fichier = f"paiements_cndk_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"

    # ── Réponse HTTP avec le fichier CSV ─────────────────────
    # BOM UTF-8 (\ufeff) pour que Excel ouvre correctement le fichier
    csv_bytes = "\ufeff" + output.getvalue()

    return Response(
        csv_bytes,
        mimetype="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename={nom_fichier}"
        }
    )


def _construire_requete_export(activite_id, classe, statut, date_debut, date_fin):
    """
    Fonction interne partagée par admin_export et admin_export_csv.
    Construit la requête SQLAlchemy selon les filtres reçus.
    Tri : activité → classe → nom élève → prénom élève
    """

    query = (
        ActiviteEleve.query
        .join(Eleve, ActiviteEleve.eleve_id == Eleve.id)
        .join(Activite, ActiviteEleve.activite_id == Activite.id)
    )

    # Filtre par activité
    if activite_id:
        query = query.filter(ActiviteEleve.activite_id == int(activite_id))

    # Filtre par classe
    if classe:
        query = query.filter(Eleve.classe == classe)

    # Filtre par statut
    if statut == "paye":
        query = query.filter(ActiviteEleve.statut == "paye")
    elif statut == "en_attente":
        query = query.filter(ActiviteEleve.statut == "en_attente")
    # "tous" → pas de filtre statut

    # Filtre par période (date de validation)
    if date_debut or date_fin:
        # On joint les paiements pour filtrer sur paye_le
        query = query.outerjoin(
            PaiementActivite,
            PaiementActivite.activite_eleve_id == ActiviteEleve.id
        )
        if date_debut:
            query = query.filter(
                db.or_(
                    PaiementActivite.paye_le >= date_debut,
                    ActiviteEleve.statut == "en_attente"  # inclut les non payés
                )
            )
        if date_fin:
            query = query.filter(
                db.or_(
                    PaiementActivite.paye_le <= date_fin + " 23:59:59",
                    ActiviteEleve.statut == "en_attente"
                )
            )

    # Tri : activité → classe → nom → prénom
    query = query.order_by(
        Activite.titre,
        Eleve.classe,
        Eleve.nom,
        Eleve.prenom
    )

    return query.all()


# ─────────────────────────────────────────────────────────────
# EXPORT BOB50 — À ajouter dans routes.py
# ─────────────────────────────────────────────────────────────
# Ajouter en haut de routes.py :
#   from app.bob50_service import generer_fichiers_bob50
#   from datetime import datetime
# ─────────────────────────────────────────────────────────────


@main.route("/admin/export/bob50")
def admin_export_bob50():
    """
    Génère et télécharge un ZIP contenant HFac.dbf + Lfac.dbf
    pour import dans Bob50.

    Seuls les paiements VALIDÉS (statut = "paye") et
    NON ENCORE EXPORTÉS (bob50_exported_at IS NULL) sont inclus.

    Après génération, bob50_exported_at est rempli sur chaque
    paiement pour éviter les doublons lors du prochain export.
    """

    # ── Récupération des paiements à exporter ────────────────
    # On joint ActiviteEleve, Eleve et Activite pour que le service
    # puisse accéder à toutes les données nécessaires
    paiements = (
        PaiementActivite.query
        .join(ActiviteEleve, PaiementActivite.activite_eleve_id == ActiviteEleve.id)
        .join(Eleve,         ActiviteEleve.eleve_id == Eleve.id)
        .join(Activite,      ActiviteEleve.activite_id == Activite.id)
        .filter(PaiementActivite.statut == "paye")
        .filter(PaiementActivite.bob50_exported_at == None)  # pas encore exportés
        .order_by(Activite.titre, Eleve.classe, Eleve.nom, Eleve.prenom)
        .all()
    )

    if not paiements:
        flash("Aucun paiement à exporter (tous déjà exportés ou aucun paiement validé).", "info")
        return redirect(url_for("main.admin_export"))

    # ── Génération des fichiers DBF ──────────────────────────
    zip_bytes = generer_fichiers_bob50(paiements)

    # ── Marquage des paiements comme exportés ────────────────
    # On remplit bob50_exported_at pour ne pas les réexporter
    maintenant = datetime.now()
    for p in paiements:
        p.bob50_exported_at = maintenant
    db.session.commit()

    # ── Nom du fichier ZIP avec date du jour ─────────────────
    nom_fichier = f"bob50_{maintenant.strftime('%Y%m%d_%H%M')}.zip"

    flash(f"{len(paiements)} paiement(s) exporté(s) vers Bob50.", "succes")

    return send_file(
        io.BytesIO(zip_bytes),
        mimetype="application/zip",
        download_name=nom_fichier,
        as_attachment=True
    )


# ─────────────────────────────────────────────────────────────
# PORTEFEUILLE ÉLÈVES
# ─────────────────────────────────────────────────────────────

@main.route("/admin/portefeuille")
def admin_portefeuille():
    """
    Page principale du portefeuille.
    Affiche la liste de tous les élèves actifs avec leur solde.
    """

    # Récupère tous les élèves actifs triés par classe puis nom
    eleves = (
        Eleve.query
        .filter(Eleve.actif == True)
        .order_by(Eleve.classe, Eleve.nom, Eleve.prenom)
        .all()
    )

    return render_template(
        "admin_portefeuille.html",
        eleves=eleves,
        role="admin",
        active="portefeuille",
        current_user_name="Administrateur (test)"
    )


@main.route("/admin/portefeuille/import", methods=["GET", "POST"])
def admin_portefeuille_import():
    """
    Import d'un extrait bancaire CSV pour recharger les portefeuilles.

    Logique d'import :
    1. Lit chaque ligne du CSV
    2. Cherche une référence structurée +++XXX/XXXX/XXXXX+++ dans toutes les cellules
    3. Extrait le matricule FASE depuis la référence
    4. Cherche l'élève correspondant dans la base
    5. Si trouvé → crédite le portefeuille + crée un MouvementPortefeuille
    6. Si non trouvé → ligne signalée comme non reconnue
    """

    ctx = {
        "role": "admin",
        "active": "portefeuille",
        "current_user_name": "Administrateur (test)"
    }

    if request.method == "GET":
        return render_template("admin_portefeuille_import.html", **ctx)

    fichier = request.files.get("fichier_csv")

    if not fichier or fichier.filename == "":
        flash("Aucun fichier sélectionné.", "erreur")
        return render_template("admin_portefeuille_import.html", **ctx)

    if not fichier.filename.lower().endswith(".csv"):
        flash("Le fichier doit être au format .csv", "erreur")
        return render_template("admin_portefeuille_import.html", **ctx)

    try:
        contenu = fichier.read().decode("utf-8-sig")
        lecteur = csv.reader(io.StringIO(contenu))

        nb_credites    = 0  # Élèves crédités avec succès
        nb_non_trouves = 0  # Références non reconnues
        nb_erreurs     = 0  # Lignes avec problèmes
        details_non_trouves = []  # Détail des lignes non reconnues

        for i, ligne in enumerate(lecteur, start=1):

            # Ignore les lignes vides
            if not any(ligne):
                continue

            # ── Recherche de la référence structurée ────────────
            # On cherche le pattern +++XXX/XXXX/XXXXX+++ dans
            # toutes les cellules de la ligne
            reference = None
            montant   = None

            for cellule in ligne:
                cellule = cellule.strip()

                # Détection référence structurée (format belge)
                # Pattern : +++XXX/XXXX/XXXXX+++
                if cellule.startswith("+++") and cellule.endswith("+++"):
                    reference = cellule
                    continue

                # Détection montant : cellule contenant un nombre décimal
                # On essaie de convertir chaque cellule en float
                if montant is None:
                    try:
                        # Remplace la virgule par un point (format européen)
                        valeur = float(cellule.replace(",", ".").replace(" ", ""))
                        if valeur > 0:
                            montant = valeur
                    except ValueError:
                        pass

            # Si pas de référence structurée trouvée → ligne ignorée
            if not reference:
                continue

            # ── Extraction du matricule FASE ─────────────────────
            # Référence : +++123/4567/89012+++
            # On extrait les chiffres : 123456789012
            matricule = reference.replace("+", "").replace("/", "").strip()

            if not matricule:
                nb_erreurs += 1
                continue

            # ── Recherche de l'élève ──────────────────────────────
            eleve = Eleve.query.filter_by(
                matricule_fase=matricule,
                actif=True
            ).first()

            if eleve is None:
                nb_non_trouves += 1
                details_non_trouves.append(
                    f"Ligne {i} — matricule {matricule} non trouvé"
                )
                continue

            # ── Crédit du portefeuille ────────────────────────────
            # Si montant non détecté automatiquement → on met 0
            # et on signale (l'admin devra vérifier)
            if montant is None:
                nb_erreurs += 1
                details_non_trouves.append(
                    f"Ligne {i} — {eleve.prenom} {eleve.nom} : montant non détecté"
                )
                continue

            # Mise à jour du solde de l'élève
            eleve.solde_portefeuille = float(eleve.solde_portefeuille) + montant

            # Création du mouvement d'historique
            mouvement = MouvementPortefeuille(
                eleve_id=eleve.id,
                montant=montant,
                type="credit",
                motif=f"Rechargement extrait bancaire — réf. {reference}"
            )
            db.session.add(mouvement)
            nb_credites += 1

        db.session.commit()

    except UnicodeDecodeError:
        flash("Erreur d'encodage. Enregistrez le fichier en UTF-8.", "erreur")
        return render_template("admin_portefeuille_import.html", **ctx)

    except Exception as e:
        db.session.rollback()
        flash(f"Erreur lors de l'import : {str(e)}", "erreur")
        return render_template("admin_portefeuille_import.html", **ctx)

    # ── Messages de résultat ──────────────────────────────────
    if nb_credites > 0:
        flash(f"{nb_credites} portefeuille(s) crédité(s) avec succès.", "succes")
    if nb_non_trouves > 0:
        flash(f"{nb_non_trouves} référence(s) non reconnue(s).", "erreur")
        for detail in details_non_trouves[:5]:
            flash(detail, "erreur")
    if nb_credites == 0 and nb_non_trouves == 0:
        flash("Aucune référence structurée trouvée dans ce fichier.", "erreur")

    return redirect(url_for("main.admin_portefeuille"))


@main.route("/admin/portefeuille/historique/<int:eleve_id>")
def admin_portefeuille_historique(eleve_id):
    """
    Affiche l'historique des mouvements du portefeuille d'un élève.
    Accessible depuis le bouton "Historique" dans admin_portefeuille.html.
    """

    # Récupère l'élève ou retourne une erreur 404 si introuvable
    eleve = Eleve.query.get_or_404(eleve_id)

    # Récupère tous les mouvements de cet élève, du plus récent au plus ancien
    mouvements = (
        MouvementPortefeuille.query
        .filter_by(eleve_id=eleve_id)
        .order_by(MouvementPortefeuille.created_at.desc())
        .all()
    )

    return render_template(
        "admin_portefeuille_historique.html",
        eleve=eleve,
        mouvements=mouvements,
        role="admin",
        active="portefeuille",
        current_user_name="Administrateur (test)"
    )


# ─────────────────────────────────────────────────────────────
# CALENDRIER REPAS — ADMIN
# ─────────────────────────────────────────────────────────────

@main.route("/admin/repas/calendrier", methods=["GET", "POST"])
def admin_calendrier_repas():
    """
    Page principale du calendrier repas.
    Affiche les types de repas et les congés encodés.
    Tous les lundi, mardi, jeudi, vendredi sont ouverts par défaut.
    Seuls les congés sont encodés ici.
    """
    from datetime import date

    types_repas = (
        TypeRepas.query
        .order_by(TypeRepas.actif.desc(), TypeRepas.nom)
        .all()
    )

    # Récupère les congés futurs triés par date
    conges = (
        CongeRepas.query
        .filter(CongeRepas.date_fin >= date.today())
        .order_by(CongeRepas.date_debut)
        .all()
    )

    return render_template(
        "admin_calendrier_repas.html",
        types_repas=types_repas,
        conges=conges,
        role="admin",
        active="calendrier_repas",
        current_user_name="Administrateur (test)"
    )


@main.route("/admin/repas/type/ajouter", methods=["POST"])
def admin_ajouter_type_repas():
    """
    Ajoute un nouveau type de repas (ex: Repas chaud, Sandwich thon).
    """
    nom         = request.form.get("nom", "").strip()
    description = request.form.get("description", "").strip() or None
    prix_str    = request.form.get("prix", "").strip()
    categorie   = request.form.get("categorie", "sandwich").strip()

    erreurs = []
    if not nom:
        erreurs.append("Le nom est obligatoire.")
    if not prix_str:
        erreurs.append("Le prix est obligatoire.")
    else:
        try:
            prix = float(prix_str)
            if prix <= 0:
                erreurs.append("Le prix doit être positif.")
        except ValueError:
            erreurs.append("Le prix doit être un nombre.")

    if erreurs:
        for e in erreurs:
            flash(e, "erreur")
        return redirect(url_for("main.admin_calendrier_repas"))

    type_repas = TypeRepas(
        nom=nom,
        description=description,
        prix=prix,
        categorie=categorie,
        actif=True
    )
    db.session.add(type_repas)
    db.session.commit()

    flash(f"Type de repas « {nom} » ajouté avec succès.", "succes")
    return redirect(url_for("main.admin_calendrier_repas"))


@main.route("/admin/repas/type/toggle/<int:type_id>", methods=["POST"])
def admin_toggle_type_repas(type_id):
    """
    Active ou désactive un type de repas.
    Un type désactivé n'apparaît plus dans les tuiles de réservation parent.
    """
    type_repas = TypeRepas.query.get_or_404(type_id)
    type_repas.actif = not type_repas.actif
    db.session.commit()

    statut = "activé" if type_repas.actif else "désactivé"
    flash(f"Type de repas « {type_repas.nom} » {statut}.", "succes")
    return redirect(url_for("main.admin_calendrier_repas"))

@main.route("/admin/repas/type/modifier/<int:type_id>", methods=["POST"])
def admin_modifier_type_repas(type_id):
    """
    Modifie le nom, la catégorie et le prix d'un type de repas.
    Le nouveau prix s'applique uniquement aux futures réservations.
    Les réservations existantes conservent leur montant d'origine.
    """
    type_repas = TypeRepas.query.get_or_404(type_id)

    nom_str   = request.form.get("nom", "").strip()
    prix_str  = request.form.get("prix", "").strip()
    categorie = request.form.get("categorie", "sandwich").strip()

    erreurs = []
    if not nom_str:
        erreurs.append("Le nom est obligatoire.")
    if not prix_str:
        erreurs.append("Le prix est obligatoire.")
    else:
        try:
            nouveau_prix = float(prix_str)
            if nouveau_prix <= 0:
                erreurs.append("Le prix doit être positif.")
        except ValueError:
            erreurs.append("Prix invalide.")

    if erreurs:
        for e in erreurs:
            flash(e, "erreur")
        return redirect(url_for("main.admin_calendrier_repas"))

    # Trace le changement de prix si modifié
    ancien_prix = float(type_repas.prix)
    type_repas.nom       = nom_str
    type_repas.prix      = nouveau_prix
    type_repas.categorie = categorie
    type_repas.updated_at = datetime.now()
    db.session.commit()

    if ancien_prix != nouveau_prix:
        flash(
            f"« {nom_str} » mis à jour — prix : "
            f"{ancien_prix:.2f}€ → {nouveau_prix:.2f}€.",
            "succes"
        )
    else:
        flash(f"« {nom_str} » mis à jour.", "succes")

    return redirect(url_for("main.admin_calendrier_repas"))

@main.route("/admin/repas/type/modifier-prix/<int:type_id>", methods=["POST"])
def admin_modifier_prix_repas(type_id):
    """
    Modifie le prix d'un type de repas.
    Le nouveau prix s'applique uniquement aux futures réservations.
    Les réservations existantes conservent leur montant d'origine.
    """
    type_repas = TypeRepas.query.get_or_404(type_id)
    prix_str   = request.form.get("nouveau_prix", "").strip()

    try:
        nouveau_prix = float(prix_str)
        if nouveau_prix <= 0:
            flash("Le prix doit être positif.", "erreur")
            return redirect(url_for("main.admin_calendrier_repas"))
    except ValueError:
        flash("Prix invalide.", "erreur")
        return redirect(url_for("main.admin_calendrier_repas"))

    ancien_prix         = float(type_repas.prix)
    type_repas.prix     = nouveau_prix
    type_repas.updated_at = datetime.now()
    db.session.commit()

    flash(
        f"Prix de « {type_repas.nom} » modifié : "
        f"{ancien_prix:.2f}€ → {nouveau_prix:.2f}€. "
        f"Les réservations existantes ne sont pas affectées.",
        "succes"
    )
    return redirect(url_for("main.admin_calendrier_repas"))


@main.route("/admin/repas/conge/ajouter", methods=["POST"])
def admin_ajouter_conge():
    """
    Ajoute une période de congé — les repas ne sont pas disponibles
    pendant cette période.
    Ex : Toussaint du 01/11 au 09/11, Noël du 23/12 au 05/01...
    """
    from datetime import date

    date_debut_str = request.form.get("date_debut", "").strip()
    date_fin_str   = request.form.get("date_fin", "").strip()
    motif          = request.form.get("motif", "").strip() or None

    if not date_debut_str or not date_fin_str:
        flash("Les deux dates sont obligatoires.", "erreur")
        return redirect(url_for("main.admin_calendrier_repas"))

    try:
        date_debut = date.fromisoformat(date_debut_str)
        date_fin   = date.fromisoformat(date_fin_str)
    except ValueError:
        flash("Format de date invalide.", "erreur")
        return redirect(url_for("main.admin_calendrier_repas"))

    if date_fin < date_debut:
        flash("La date de fin doit être après la date de début.", "erreur")
        return redirect(url_for("main.admin_calendrier_repas"))

    conge = CongeRepas(
        date_debut=date_debut,
        date_fin=date_fin,
        motif=motif
    )
    db.session.add(conge)
    db.session.commit()

    flash(
        f"Congé « {motif or 'sans motif'} » ajouté "
        f"du {date_debut.strftime('%d/%m/%Y')} "
        f"au {date_fin.strftime('%d/%m/%Y')}.",
        "succes"
    )
    return redirect(url_for("main.admin_calendrier_repas"))


@main.route("/admin/repas/conge/supprimer/<int:conge_id>", methods=["POST"])
def admin_supprimer_conge(conge_id):
    """
    Supprime un congé — les jours redeviennent disponibles
    pour les réservations.
    """
    conge = CongeRepas.query.get_or_404(conge_id)
    motif = conge.motif or "sans motif"
    db.session.delete(conge)
    db.session.commit()

    flash(f"Congé « {motif} » supprimé.", "succes")
    return redirect(url_for("main.admin_calendrier_repas"))

@main.route("/admin/repas/conge/importer", methods=["POST"])
def admin_importer_conges_csv():
    """
    Import CSV des congés scolaires.
    Format attendu : motif,date_debut,date_fin
    Les dates doivent être au format YYYY-MM-DD (ex: 2025-11-01).
    Les congés déjà existants (même date_debut) ne sont pas dupliqués.
    """
    from datetime import date

    fichier = request.files.get("fichier_csv")

    if not fichier or fichier.filename == "":
        flash("Aucun fichier sélectionné.", "erreur")
        return redirect(url_for("main.admin_calendrier_repas"))

    if not fichier.filename.lower().endswith(".csv"):
        flash("Le fichier doit être au format .csv", "erreur")
        return redirect(url_for("main.admin_calendrier_repas"))

    try:
        contenu = fichier.read().decode("utf-8-sig")
        lecteur = csv.DictReader(io.StringIO(contenu))

        # Vérifie que les colonnes requises sont présentes
        colonnes_requises = {"date_debut", "date_fin"}
        if not colonnes_requises.issubset(set(lecteur.fieldnames or [])):
            flash(
                "Colonnes manquantes. Format attendu : motif,date_debut,date_fin",
                "erreur"
            )
            return redirect(url_for("main.admin_calendrier_repas"))

        # Récupère les dates_debut déjà existantes pour éviter les doublons
        dates_existantes = {
            c.date_debut for c in CongeRepas.query.all()
        }

        nb_ajoutes  = 0
        nb_ignores  = 0
        nb_erreurs  = 0

        for i, ligne in enumerate(lecteur, start=2):
            date_debut_str = ligne.get("date_debut", "").strip()
            date_fin_str   = ligne.get("date_fin", "").strip()
            motif          = ligne.get("motif", "").strip() or None

            if not date_debut_str or not date_fin_str:
                nb_erreurs += 1
                continue

            try:
                date_debut = date.fromisoformat(date_debut_str)
                date_fin   = date.fromisoformat(date_fin_str)
            except ValueError:
                flash(
                    f"Ligne {i} ignorée — format de date invalide : "
                    f"{date_debut_str} / {date_fin_str}. "
                    "Utilisez le format YYYY-MM-DD.",
                    "erreur"
                )
                nb_erreurs += 1
                continue

            if date_fin < date_debut:
                nb_erreurs += 1
                continue

            # Ignore si ce congé existe déjà
            if date_debut in dates_existantes:
                nb_ignores += 1
                continue

            conge = CongeRepas(
                date_debut=date_debut,
                date_fin=date_fin,
                motif=motif
            )
            db.session.add(conge)
            dates_existantes.add(date_debut)
            nb_ajoutes += 1

        db.session.commit()

    except UnicodeDecodeError:
        flash("Erreur d'encodage. Enregistrez le fichier en UTF-8.", "erreur")
        return redirect(url_for("main.admin_calendrier_repas"))

    except Exception as e:
        db.session.rollback()
        flash(f"Erreur lors de l'import : {str(e)}", "erreur")
        return redirect(url_for("main.admin_calendrier_repas"))

    if nb_ajoutes > 0:
        flash(f"{nb_ajoutes} congé(s) importé(s) avec succès.", "succes")
    if nb_ignores > 0:
        flash(f"{nb_ignores} congé(s) ignoré(s) — déjà présents.", "info")
    if nb_erreurs > 0:
        flash(f"{nb_erreurs} ligne(s) ignorée(s) — erreur de format.", "erreur")

    return redirect(url_for("main.admin_calendrier_repas"))


    # ─────────────────────────────────────────────────────────────
# RÉSERVATION REPAS — PARENT
# ─────────────────────────────────────────────────────────────

@main.route("/parent/repas")
def parent_reservation_repas():
    """
    Calendrier mensuel de réservation des repas.
    Affiche le mois en cours par défaut, navigation possible.
    Chaque jour cliquable ouvre un pop-up avec les tuiles de repas.
    Réservation bloquée après 9h00 le matin même.
    """
    from datetime import date, datetime
    import calendar

    parent_id = 1  # TODO : flask_login.current_user.id

    # ── Mois affiché — paramètre GET ou mois en cours ────────
    # ?mois=2026-05 → affiche mai 2026
    mois_str = request.args.get("mois", "")
    try:
        if mois_str:
            annee, mois = map(int, mois_str.split("-"))
        else:
            annee = date.today().year
            mois  = date.today().month
    except ValueError:
        annee = date.today().year
        mois  = date.today().month

    # ── Navigation mois précédent / suivant ──────────────────
    if mois == 1:
        mois_precedent = f"{annee - 1}-12"
    else:
        mois_precedent = f"{annee}-{mois - 1:02d}"

    if mois == 12:
        mois_suivant = f"{annee + 1}-01"
    else:
        mois_suivant = f"{annee}-{mois + 1:02d}"

    # ── Génération des jours du mois ─────────────────────────
    # calendar.monthcalendar retourne une liste de semaines
    # Chaque semaine est une liste de 7 jours (0 = jour hors mois)
    semaines = calendar.monthcalendar(annee, mois)

    # ── Récupération des congés du mois ──────────────────────
    # Pour savoir quels jours sont fermés
    premier_jour = date(annee, mois, 1)
    dernier_jour = date(annee, mois, calendar.monthrange(annee, mois)[1])

    conges = CongeRepas.query.filter(
        CongeRepas.date_debut <= dernier_jour,
        CongeRepas.date_fin   >= premier_jour
    ).all()

    # Construit un set de toutes les dates en congé pour lookup rapide
    dates_conge = set()
    for c in conges:
        d = c.date_debut
        while d <= c.date_fin:
            dates_conge.add(d)
            from datetime import timedelta
            d += timedelta(days=1)

    # ── Récupération des réservations du parent ce mois ──────
    # On récupère les enfants du parent d'abord
    eleves_parent = (
        Eleve.query
        .join(ParentEleve, ParentEleve.eleve_id == Eleve.id)
        .filter(ParentEleve.parent_id == parent_id)
        .filter(Eleve.actif == True)
        .all()
    )

    eleve_ids = [e.id for e in eleves_parent]

    # Réservations confirmées ce mois pour les enfants du parent
    reservations = (
        ReservationRepas.query
        .filter(ReservationRepas.eleve_id.in_(eleve_ids))
        .filter(ReservationRepas.date_repas >= premier_jour)
        .filter(ReservationRepas.date_repas <= dernier_jour)
        .filter(ReservationRepas.statut == "confirme")
        .all()
    )

    # Construit un dict {(eleve_id, date): reservation} pour lookup rapide
    reservations_par_jour = {
        (r.eleve_id, r.date_repas): r for r in reservations
    }

    # ── Types de repas actifs ─────────────────────────────────
    types_repas = (
        TypeRepas.query
        .filter(TypeRepas.actif == True)
        .order_by(TypeRepas.categorie.desc(), TypeRepas.nom)
        .all()
    )

    # ── Heure limite poru passer une commande : 9h00 le matin même ────────────────────
    maintenant = datetime.now()
    heure_limite_passee = maintenant.hour >= 9

    return render_template(
        "reservation_repas.html",
        annee=annee,
        mois=mois,
        mois_precedent=mois_precedent,
        mois_suivant=mois_suivant,
        semaines=semaines,
        dates_conge=dates_conge,
        eleves_parent=eleves_parent,
        reservations_par_jour=reservations_par_jour,
        types_repas=types_repas,
        heure_limite_passee=heure_limite_passee,
        maintenant=maintenant,
        role="parent",
        active="repas",
        current_user_name="Parent (test)"
    )


@main.route("/parent/repas/reserver", methods=["POST"])
def parent_reserver_repas():
    """
    Enregistre une réservation de repas pour un élève.
    Vérifie :
    - L'heure limite (9h00 le matin même)
    - Que l'élève appartient bien au parent connecté
    - Que le jour n'est pas un congé
    - Que le solde est suffisant
    - Qu'il n'y a pas déjà une réservation ce jour
    """
    from datetime import date, datetime

    parent_id = 1  # TODO : flask_login.current_user.id

    eleve_id      = request.form.get("eleve_id", type=int)
    type_repas_id = request.form.get("type_repas_id", type=int)
    date_str      = request.form.get("date_repas", "").strip()
    mois_str      = request.form.get("mois", "")

    # ── Validation de base ────────────────────────────────────
    if not all([eleve_id, type_repas_id, date_str]):
        flash("Données manquantes.", "erreur")
        return redirect(url_for("main.parent_reservation_repas", mois=mois_str))

    try:
        date_repas = date.fromisoformat(date_str)
    except ValueError:
        flash("Date invalide.", "erreur")
        return redirect(url_for("main.parent_reservation_repas", mois=mois_str))

    # ── Vérification heure limite ─────────────────────────────
    maintenant = datetime.now()
    if date_repas == date.today() and maintenant.hour >= 9:
        flash("Les réservations sont closes depuis 9h00 ce matin.", "erreur")
        return redirect(url_for("main.parent_reservation_repas", mois=mois_str))

    if date_repas < date.today():
        flash("Impossible de réserver pour une date passée.", "erreur")
        return redirect(url_for("main.parent_reservation_repas", mois=mois_str))

    # ── Vérification que l'élève appartient au parent ─────────
    lien = ParentEleve.query.filter_by(
        parent_id=parent_id,
        eleve_id=eleve_id
    ).first()

    if not lien:
        flash("Accès non autorisé.", "erreur")
        return redirect(url_for("main.parent_reservation_repas", mois=mois_str))

    eleve      = Eleve.query.get_or_404(eleve_id)
    type_repas = TypeRepas.query.get_or_404(type_repas_id)

    # ── Vérification congé ────────────────────────────────────
    conge = CongeRepas.query.filter(
        CongeRepas.date_debut <= date_repas,
        CongeRepas.date_fin   >= date_repas
    ).first()

    if conge:
        flash("Aucun repas disponible ce jour (congé scolaire).", "erreur")
        return redirect(url_for("main.parent_reservation_repas", mois=mois_str))

    # ── Vérification doublon ──────────────────────────────────
    existe = ReservationRepas.query.filter_by(
        eleve_id=eleve_id,
        date_repas=date_repas,
        statut="confirme"
    ).first()

    if existe:
        flash(f"{eleve.prenom} a déjà un repas réservé ce jour.", "erreur")
        return redirect(url_for("main.parent_reservation_repas", mois=mois_str))

    # ── Vérification solde suffisant ──────────────────────────
    if float(eleve.solde_portefeuille) < float(type_repas.prix):
        flash(
            f"Solde insuffisant pour {eleve.prenom} "
            f"(solde : {float(eleve.solde_portefeuille):.2f}€ / "
            f"prix : {float(type_repas.prix):.2f}€).",
            "erreur"
        )
        return redirect(url_for("main.parent_reservation_repas", mois=mois_str))

    # ── Création de la réservation ────────────────────────────
    reservation = ReservationRepas(
        eleve_id=eleve_id,
        type_repas_id=type_repas_id,
        date_repas=date_repas,
        montant=type_repas.prix,
        statut="confirme"
    )
    db.session.add(reservation)

    # ── Débit du portefeuille ─────────────────────────────────
    eleve.solde_portefeuille = float(eleve.solde_portefeuille) - float(type_repas.prix)

    mouvement = MouvementPortefeuille(
        eleve_id=eleve_id,
        montant=type_repas.prix,
        type="debit",
        motif=f"Réservation {type_repas.nom} — {date_repas.strftime('%d/%m/%Y')}"
    )
    db.session.add(mouvement)
    db.session.commit()

    flash(
        f"Repas « {type_repas.nom} » réservé pour {eleve.prenom} "
        f"le {date_repas.strftime('%d/%m/%Y')} "
        f"({float(type_repas.prix):.2f}€ débité).",
        "succes"
    )
    return redirect(url_for("main.parent_reservation_repas", mois=mois_str))


@main.route("/parent/repas/annuler/<int:reservation_id>", methods=["POST"])
def parent_annuler_repas(reservation_id):
    """
    Annule une réservation de repas et rembourse le portefeuille.
    Annulation possible jusqu'à 9h00 le matin même.
    """
    from datetime import date, datetime

    parent_id = 1  # TODO : flask_login.current_user.id

    reservation = ReservationRepas.query.get_or_404(reservation_id)
    mois_str = request.form.get("mois", "")

    # ── Vérification que l'élève appartient au parent ─────────
    lien = ParentEleve.query.filter_by(
        parent_id=parent_id,
        eleve_id=reservation.eleve_id
    ).first()

    if not lien:
        flash("Accès non autorisé.", "erreur")
        return redirect(url_for("main.parent_reservation_repas", mois=mois_str))

    # ── Vérification heure limite ─────────────────────────────
    maintenant = datetime.now()
    if reservation.date_repas == date.today() and maintenant.hour >= 9:
        flash("Les annulations sont closes depuis 9h00 ce matin.", "erreur")
        return redirect(url_for("main.parent_reservation_repas", mois=mois_str))

    if reservation.date_repas < date.today():
        flash("Impossible d'annuler une réservation passée.", "erreur")
        return redirect(url_for("main.parent_reservation_repas", mois=mois_str))

    # ── Annulation + remboursement ────────────────────────────
    eleve = Eleve.query.get(reservation.eleve_id)

    reservation.statut = "annule"

    # Remboursement du portefeuille
    eleve.solde_portefeuille = float(eleve.solde_portefeuille) + float(reservation.montant)

    mouvement = MouvementPortefeuille(
        eleve_id=eleve.id,
        montant=reservation.montant,
        type="remb",
        motif=f"Annulation {reservation.type_repas.nom} — "
              f"{reservation.date_repas.strftime('%d/%m/%Y')}"
    )
    db.session.add(mouvement)
    db.session.commit()

    flash(
        f"Réservation annulée — {float(reservation.montant):.2f}€ "
        f"remboursé sur le portefeuille de {eleve.prenom}.",
        "succes"
    )
    return redirect(url_for("main.parent_reservation_repas", mois=mois_str))


@main.route("/parent/repas/historique")
def parent_historique_repas():
    """
    Historique des repas réservés pour les enfants du parent connecté.
    Triés du plus récent au plus ancien.
    """
    parent_id = 1  # TODO : flask_login.current_user.id

    eleve_ids = [
        lien.eleve_id for lien in
        ParentEleve.query.filter_by(parent_id=parent_id).all()
    ]

    reservations = (
        ReservationRepas.query
        .filter(ReservationRepas.eleve_id.in_(eleve_ids))
        .order_by(ReservationRepas.date_repas.desc())
        .all()
    )

    return render_template(
        "historique_repas.html",
        reservations=reservations,
        role="parent",
        active="repas",
        current_user_name="Parent (test)"
    )